# scripts/excel_export.py
import sys
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from datetime import datetime

def get_template_url(trainee_count):
    """Determine which template to use based on trainee count"""
    if trainee_count > 250:
        return 'Directory-300.xlsx'
    elif trainee_count > 200:
        return 'Directory-250.xlsx'
    elif trainee_count > 150:
        return 'Directory-200.xlsx'
    elif trainee_count > 100:
        return 'Directory-150.xlsx'
    elif trainee_count > 50:
        return 'Directory-100.xlsx'
    elif trainee_count > 30:
        return 'Directory-0.xlsx'
    elif trainee_count > 25:
        return 'Directory-1.xlsx'
    elif trainee_count > 20:
        return 'Directory-2.xlsx'
    elif trainee_count > 15:
        return 'Directory-3.xlsx'
    elif trainee_count > 10:
        return 'Directory-4.xlsx'
    else:
        return 'Directory-5.xlsx'

def main():
    # Read JSON data from stdin
    input_data = json.loads(sys.stdin.read())
    
    trainees = input_data['trainees']
    course_name = input_data['courseName']
    training_dates = input_data['trainingDates']
    schedule_id = input_data['scheduleId']
    template_dir = input_data['templateDir']
    output_path = input_data['outputPath']
    proxy_url = input_data.get('proxyUrl', '')
    
    trainee_count = len(trainees)
    
    # Load the appropriate template
    template_file = get_template_url(trainee_count)
    template_path = f"{template_dir}/{template_file}"
    
    print(f"Loading template: {template_path} for {trainee_count} trainees", file=sys.stderr)
    
    wb = load_workbook(template_path)
    ws = wb['Directory of Participants']
    
    # Fill header information (rows 10-11)
    ws['C10'] = course_name
    ws['C11'] = training_dates
    
    # Count gender
    male_count = sum(1 for t in trainees if t.get('gender', '').lower() == 'male')
    female_count = sum(1 for t in trainees if t.get('gender', '').lower() == 'female')
    
    # Participant rows start from row 15
    start_row = 15
    
    for i, trainee in enumerate(trainees):
        row_num = start_row + i
        
        # Fill data
        ws[f'A{row_num}'] = i + 1
        # ws[f'B{row_num}'] = trainee.get('certificate_number', '')  # Commented out like PHP
        ws[f'C{row_num}'] = (trainee.get('last_name', '') or '').upper()
        ws[f'D{row_num}'] = (trainee.get('first_name', '') or '').upper()
        
        middle_initial = trainee.get('middle_initial', '')
        ws[f'E{row_num}'] = middle_initial[0].upper() if middle_initial else ''
        
        ws[f'F{row_num}'] = (trainee.get('suffix', '') or '').upper()
        ws[f'G{row_num}'] = trainee.get('gender', '')
        ws[f'H{row_num}'] = trainee.get('age', '')
        ws[f'I{row_num}'] = trainee.get('company_name', '')
        ws[f'J{row_num}'] = trainee.get('company_position', '')
        ws[f'K{row_num}'] = trainee.get('company_city', '')
        ws[f'L{row_num}'] = trainee.get('company_region', '')
        ws[f'M{row_num}'] = trainee.get('company_industry', '')
        ws[f'N{row_num}'] = trainee.get('total_workers', '')
        ws[f'O{row_num}'] = trainee.get('company_email', '')
        ws[f'P{row_num}'] = trainee.get('email', '')
        ws[f'Q{row_num}'] = trainee.get('phone_number', '')
        ws[f'R{row_num}'] = trainee.get('company_landline', '')
        ws[f'T{row_num}'] = 'Online Training'
        ws[f'U{row_num}'] = f"#{trainee.get('schedule_id', '')[-4:]}"
        
        # Add image if picture URL exists
        picture_url = trainee.get('picture_2x2_url', '')
        if picture_url:
            try:
                # Use proxy URL if provided
                if proxy_url:
                    img_url = f"{proxy_url}?url={picture_url}"
                else:
                    img_url = picture_url
                
                print(f"Fetching image {i+1}: {img_url}", file=sys.stderr)
                
                response = requests.get(img_url, timeout=10)
                if response.status_code == 200:
                    img = Image(BytesIO(response.content))
                    
                    # Set image size to 96x96 pixels (matching PHP)
                    img.width = 96
                    img.height = 96
                    
                    # Add image to cell S (column 19)
                    ws.add_image(img, f'S{row_num}')
                    print(f"✓ Image {i+1} added successfully", file=sys.stderr)
                else:
                    print(f"✗ Failed to fetch image {i+1}: HTTP {response.status_code}", file=sys.stderr)
            except Exception as e:
                print(f"✗ Error adding image {i+1}: {str(e)}", file=sys.stderr)
    
    # Update Training Database sheet if it exists
    try:
        db_ws = wb['Training Database']
        db_ws['A14'] = 1
        db_ws['B14'] = course_name
        db_ws['C14'] = training_dates
        db_ws['D14'] = f"#{schedule_id[-4:]}"
        db_ws['E14'] = trainee_count
        db_ws['F14'] = male_count
        db_ws['G14'] = female_count
        db_ws['M14'] = 'Online'
    except:
        print("Training Database sheet not found, skipping...", file=sys.stderr)
    
    # Save the workbook
    wb.save(output_path)
    print(f"✓ Excel file saved: {output_path}", file=sys.stderr)
    
    # Output success to stdout
    print(json.dumps({"success": True, "path": output_path}))

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)